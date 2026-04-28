import os
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from sqlalchemy import select, func
from datetime import datetime, timedelta
from app.database import async_session, Student, Transaction, Attendance, AttendanceStatus
from app.bot.states import TransactionState, AttendanceState, InfoState, EditBalanceState
from app.bot import keyboards

import asyncio
from aiogram.filters import Command

router = Router()

LESSON_PRICE = float(os.getenv("LESSON_PRICE", 250))
ADMIN_ID = os.getenv("ADMIN_ID")
ASSISTANT_ID = os.getenv("ASSISTANT_ID")

@router.message.outer_middleware()
async def access_middleware(handler, event, data):
    user = data.get("event_from_user")
    if user:
        allowed_ids = [str(ADMIN_ID), str(ASSISTANT_ID)]
        if str(user.id) not in allowed_ids:
            await event.answer("У вас немає доступу до цього бота.")
            return
    return await handler(event, data)

@router.callback_query.outer_middleware()
async def access_middleware_cb(handler, event, data):
    user = data.get("event_from_user")
    if user:
        allowed_ids = [str(ADMIN_ID), str(ASSISTANT_ID)]
        if str(user.id) not in allowed_ids:
            await event.answer("У вас немає доступу.", show_alert=True)
            return
    return await handler(event, data)

@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Привіт! Я бот для моніторингу занять.", reply_markup=keyboards.get_main_keyboard(message.from_user.id))

@router.message(F.text == "Скасувати")
async def cmd_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Дію скасовано.", reply_markup=keyboards.get_main_keyboard(message.from_user.id))

# --- ВНЕСЕННЯ ТРАНЗАКЦІЇ ---
@router.message(F.text == "Внести транзакцію")
async def process_transaction_start(message: Message, state: FSMContext):
    async with async_session() as session:
        result = await session.execute(select(Student))
        students = result.scalars().all()
    
    if not students:
        await message.answer("Учнів ще немає в базі. Напишіть Прізвище та Ім'я нового учня:", reply_markup=keyboards.get_cancel_keyboard())
    else:
        await message.answer("Виберіть учня або напишіть Прізвище та Ім'я нового:", reply_markup=keyboards.get_students_inline_keyboard(students))
    
    await state.set_state(TransactionState.waiting_for_student)

@router.callback_query(TransactionState.waiting_for_student, F.data.startswith("student_"))
async def process_transaction_student_cb(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[1])
    await state.update_data(student_id=student_id)
    await callback.message.answer("Введіть суму надходження (число):", reply_markup=keyboards.get_cancel_keyboard())
    await state.set_state(TransactionState.waiting_for_amount)
    await callback.answer()

@router.message(TransactionState.waiting_for_student)
async def process_transaction_new_student(message: Message, state: FSMContext):
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Будь ласка, введіть Прізвище та Ім'я (два слова).")
        return
    
    async with async_session() as session:
        new_student = Student(last_name=parts[0], first_name=parts[1])
        session.add(new_student)
        await session.commit()
        await session.refresh(new_student)
        await state.update_data(student_id=new_student.id)
        
    await message.answer(f"Створено нового учня: {new_student.last_name} {new_student.first_name}.\nВведіть суму надходження (число):", reply_markup=keyboards.get_cancel_keyboard())
    await state.set_state(TransactionState.waiting_for_amount)

@router.message(TransactionState.waiting_for_amount)
async def process_transaction_amount(message: Message, state: FSMContext):
    try:
        amount = int(message.text)
    except ValueError:
        await message.answer("Будь ласка, введіть коректне число.")
        return
    
    await state.update_data(amount=amount)
    
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сьогодні")],
            [KeyboardButton(text="Вчора")],
            [KeyboardButton(text="Скасувати")]
        ],
        resize_keyboard=True
    )
    await message.answer("Введіть дату транзакції у форматі ДД.ММ.РРРР або оберіть швидкий варіант:", reply_markup=kb)
    await state.set_state(TransactionState.waiting_for_date)

@router.message(TransactionState.waiting_for_date)
async def process_transaction_date(message: Message, state: FSMContext):
    text = message.text
    if text == "Сьогодні":
        date = datetime.now()
    elif text == "Вчора":
        date = datetime.now() - timedelta(days=1)
    else:
        try:
            date = datetime.strptime(text, "%d.%m.%Y")
        except ValueError:
            await message.answer("Некоректний формат дати. Введіть у форматі ДД.ММ.РРРР (наприклад 25.10.2023) або оберіть 'Сьогодні'/'Вчора'.")
            return
            
    await state.update_data(date=date)
    
    data = await state.get_data()
    student_id = data["student_id"]
    amount = data["amount"]
    lessons_added = amount / LESSON_PRICE
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        student_name = f"{student.last_name} {student.first_name}"
        await state.update_data(student_name=student_name, lessons_added=lessons_added)

    summary = (
        f"📋 Підтвердіть транзакцію:\n"
        f"👤 Учень: {student_name}\n"
        f"💰 Сума: {amount} грн\n"
        f"📅 Дата: {date.strftime('%d.%m.%Y')}\n"
        f"📚 Буде додано занять: {lessons_added}\n"
    )
    
    await message.answer(summary, reply_markup=keyboards.get_confirmation_keyboard())
    await state.set_state(TransactionState.waiting_for_confirmation)

@router.message(TransactionState.waiting_for_confirmation, F.text == "Так, все вірно")
async def process_transaction_confirmation(message: Message, state: FSMContext):
    data = await state.get_data()
    student_id = data["student_id"]
    amount = data["amount"]
    date = data["date"]
    lessons_added = data["lessons_added"]
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        
        transaction = Transaction(student_id=student_id, date=date, amount=amount, lessons_added=lessons_added)
        session.add(transaction)
        
        student.balance_lessons += lessons_added
        await session.commit()
        new_balance = student.balance_lessons
        
    await message.answer(
        f"✅ Транзакцію збережено!\nНовий баланс: {new_balance} занять", 
        reply_markup=keyboards.get_main_keyboard(message.from_user.id)
    )
    await state.clear()

# --- ВНЕСЕННЯ ВІДВІДУВАННЯ ---
@router.message(F.text == "Внести відвідування")
async def process_attendance_start(message: Message, state: FSMContext):
    async with async_session() as session:
        result = await session.execute(select(Student))
        students = result.scalars().all()
    
    if not students:
        await message.answer("Учнів ще немає в базі.")
        return
        
    await message.answer("Виберіть учня:", reply_markup=keyboards.get_students_inline_keyboard(students))
    await state.set_state(AttendanceState.waiting_for_student)

@router.callback_query(AttendanceState.waiting_for_student, F.data.startswith("student_"))
async def process_attendance_student_cb(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[1])
    await state.update_data(student_id=student_id)
    await callback.message.answer("Оберіть статус відвідування:", reply_markup=keyboards.get_attendance_status_keyboard())
    await state.set_state(AttendanceState.waiting_for_status)
    await callback.answer()

@router.message(AttendanceState.waiting_for_status)
async def process_attendance_status(message: Message, state: FSMContext):
    status_map = {
        "Була (-1 урок)": (AttendanceStatus.PRESENT, -1),
        "Не була (перенесено)": (AttendanceStatus.ABSENT_RESCHEDULED, 0),
        "Не була (списано)": (AttendanceStatus.ABSENT_CHARGED, -1)
    }
    
    if message.text not in status_map:
        await message.answer("Будь ласка, оберіть статус з клавіатури.")
        return
        
    db_status, balance_change = status_map[message.text]
    await state.update_data(db_status=db_status, balance_change=balance_change, status_text=message.text)
    
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сьогодні")],
            [KeyboardButton(text="Вчора")],
            [KeyboardButton(text="Скасувати")]
        ],
        resize_keyboard=True
    )
    await message.answer("Введіть дату відвідування у форматі ДД.ММ.РРРР або оберіть швидкий варіант:", reply_markup=kb)
    await state.set_state(AttendanceState.waiting_for_date)

@router.message(AttendanceState.waiting_for_date)
async def process_attendance_date(message: Message, state: FSMContext):
    text = message.text
    if text == "Сьогодні":
        date = datetime.now()
    elif text == "Вчора":
        date = datetime.now() - timedelta(days=1)
    else:
        try:
            date = datetime.strptime(text, "%d.%m.%Y")
        except ValueError:
            await message.answer("Некоректний формат дати. Введіть у форматі ДД.ММ.РРРР (наприклад 25.10.2023) або оберіть 'Сьогодні'/'Вчора'.")
            return
            
    await state.update_data(date=date)
    
    data = await state.get_data()
    student_id = data["student_id"]
    status_text = data["status_text"]
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        student_name = f"{student.last_name} {student.first_name}"
        await state.update_data(student_name=student_name)

    summary = (
        f"📋 Підтвердіть відвідування:\n"
        f"👤 Учень: {student_name}\n"
        f"📝 Статус: {status_text}\n"
        f"📅 Дата: {date.strftime('%d.%m.%Y')}\n"
    )
    
    await message.answer(summary, reply_markup=keyboards.get_confirmation_keyboard())
    await state.set_state(AttendanceState.waiting_for_confirmation)

@router.message(AttendanceState.waiting_for_confirmation, F.text == "Так, все вірно")
async def process_attendance_confirmation(message: Message, state: FSMContext):
    data = await state.get_data()
    student_id = data["student_id"]
    db_status = data["db_status"]
    balance_change = data["balance_change"]
    date = data["date"]
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        
        # Перевірка на наявність запису на ту ж дату (ігноруючи час)
        existing_query = select(Attendance).where(
            Attendance.student_id == student_id,
            func.date(Attendance.date) == date.date()
        )
        existing_res = await session.execute(existing_query)
        existing = existing_res.scalars().first() # Беремо перший знайдений запис
        
        if existing:
            # Відкочуємо попередній вплив і додаємо новий
            student.balance_lessons -= existing.balance_impact
            student.balance_lessons += balance_change
            
            existing.status = db_status
            existing.balance_impact = balance_change
            existing.date = date # Оновлюємо точний час якщо треба
            msg_text = "🔄 Відвідування на цю дату вже було, дані перезаписано!"
        else:
            attendance = Attendance(student_id=student_id, date=date, status=db_status, balance_impact=balance_change)
            session.add(attendance)
            student.balance_lessons += balance_change
            msg_text = "✅ Відвідування збережено!"
            
        await session.commit()
        new_balance = student.balance_lessons
        
    await message.answer(
        f"{msg_text}\nНовий баланс: {new_balance} занять", 
        reply_markup=keyboards.get_main_keyboard(message.from_user.id)
    )
    await state.clear()

# --- РЕДАГУВАННЯ БАЛАНСУ ---
@router.message(F.text == "Редагувати баланс")
async def process_edit_balance_start(message: Message, state: FSMContext):
    async with async_session() as session:
        result = await session.execute(select(Student))
        students = result.scalars().all()
    
    if not students:
        await message.answer("Учнів ще немає в базі.")
        return
        
    await message.answer("Виберіть учня для редагування балансу:", reply_markup=keyboards.get_students_inline_keyboard(students))
    await state.set_state(EditBalanceState.waiting_for_student)

@router.callback_query(EditBalanceState.waiting_for_student, F.data.startswith("student_"))
async def process_edit_balance_student_cb(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[1])
    async with async_session() as session:
        student = await session.get(Student, student_id)
    
    await state.update_data(student_id=student_id, student_name=f"{student.last_name} {student.first_name}")
    await callback.message.answer(
        f"👤 Учень: {student.last_name} {student.first_name}\n"
        f"💰 Поточний баланс: {student.balance_lessons} занять\n\n"
        f"Введіть число для зміни балансу (наприклад, +2 або -1.5):",
        reply_markup=keyboards.get_cancel_keyboard()
    )
    await state.set_state(EditBalanceState.waiting_for_amount)
    await callback.answer()

@router.message(EditBalanceState.waiting_for_amount)
async def process_edit_balance_amount(message: Message, state: FSMContext):
    try:
        amount_str = message.text.replace(",", ".")
        amount = float(amount_str)
    except ValueError:
        await message.answer("Будь ласка, введіть коректне число (наприклад +1 або -1).")
        return
    
    data = await state.get_data()
    student_name = data["student_name"]
    await state.update_data(amount=amount)
    
    summary = (
        f"📋 Підтвердіть зміну балансу:\n"
        f"👤 Учень: {student_name}\n"
        f"🔢 Зміна: {'+' if amount > 0 else ''}{amount} занять"
    )
    
    await message.answer(summary, reply_markup=keyboards.get_confirmation_keyboard())
    await state.set_state(EditBalanceState.waiting_for_confirmation)

@router.message(EditBalanceState.waiting_for_confirmation, F.text == "Так, все вірно")
async def process_edit_balance_confirmation(message: Message, state: FSMContext):
    data = await state.get_data()
    student_id = data["student_id"]
    amount = data["amount"]
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        student.balance_lessons += amount
        await session.commit()
        new_balance = student.balance_lessons
        
    await message.answer(
        f"✅ Баланс відредаговано!\nНовий баланс: {new_balance} занять", 
        reply_markup=keyboards.get_main_keyboard(message.from_user.id)
    )
    await state.clear()

# --- ІНФОРМАЦІЯ ПО УЧНЮ ---
@router.message(F.text == "Інформація по учню")
async def process_info_start(message: Message, state: FSMContext):
    async with async_session() as session:
        result = await session.execute(select(Student))
        students = result.scalars().all()
    
    if not students:
        await message.answer("Учнів ще немає в базі.")
        return
        
    await message.answer("Виберіть учня:", reply_markup=keyboards.get_students_inline_keyboard(students))
    await state.set_state(InfoState.waiting_for_student)

@router.callback_query(InfoState.waiting_for_student, F.data.startswith("student_"))
async def process_info_student_cb(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[1])
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        
        # Отримуємо останні 5 відвідувань
        att_res = await session.execute(select(Attendance).where(Attendance.student_id == student_id).order_by(Attendance.date.desc()).limit(5))
        attendances = att_res.scalars().all()
        
    text = f"👤 Учень: {student.last_name} {student.first_name}\n"
    text += f"💰 Баланс: {student.balance_lessons} занять\n\n"
    
    text += "📅 Останні відвідування:\n"
    for att in attendances:
        status_str = {
            AttendanceStatus.PRESENT: "Була",
            AttendanceStatus.ABSENT_RESCHEDULED: "Не була (перенесено)",
            AttendanceStatus.ABSENT_CHARGED: "Не була (списано)"
        }.get(att.status, "Невідомо")
        text += f"- {att.date.strftime('%Y-%m-%d %H:%M')}: {status_str}\n"
        
    if not attendances:
        text += "- Немає записів\n"
        
    # Створюємо кнопку для Excel
    excel_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📊 Завантажити повний звіт (Excel)", callback_data=f"excel_{student_id}")]
    ])
    
    # Відправляємо ОДНЕ повідомлення з даними та кнопкою
    await callback.message.answer(text, reply_markup=excel_kb)
    
    await state.clear()
    await callback.answer()

from aiogram.types import FSInputFile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@router.callback_query(F.data.startswith("excel_"))
async def process_excel_export(callback: CallbackQuery):
    student_id = int(callback.data.split("_")[1])
    
    async with async_session() as session:
        student = await session.get(Student, student_id)
        
        # Отримуємо всі дані
        trans_res = await session.execute(select(Transaction).where(Transaction.student_id == student_id).order_by(Transaction.date.asc()))
        transactions = trans_res.scalars().all()
        
        att_res = await session.execute(select(Attendance).where(Attendance.student_id == student_id).order_by(Attendance.date.asc()))
        attendances = att_res.scalars().all()

    # Створюємо Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Звіт"
    
    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Звіт по учню: {student.last_name} {student.first_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Поточний баланс: {student.balance_lessons} занять"
    ws['A2'].font = Font(italic=True)
    
    # Шапка таблиці
    headers = ["Дата", "Тип події", "Деталі", "Зміна балансу", "Сума (грн)"]
    ws.append([])
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True)
    
    # Збираємо всі події в один список для сортування
    events = []
    for t in transactions:
        events.append({
            'date': t.date,
            'type': 'Оплата',
            'details': f"Зараховано {t.lessons_added} занять",
            'impact': t.lessons_added,
            'amount': t.amount
        })
    
    for a in attendances:
        status_str = {
            AttendanceStatus.PRESENT: "Присутність",
            AttendanceStatus.ABSENT_RESCHEDULED: "Не була (перенесено)",
            AttendanceStatus.ABSENT_CHARGED: "Не була (списано)"
        }.get(a.status, "Невідомо")
        
        events.append({
            'date': a.date,
            'type': 'Заняття',
            'details': status_str,
            'impact': a.balance_impact,
            'amount': "-"
        })
        
    # Стиль меж
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    events.sort(key=lambda x: x['date'])
    
    # Кольори
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Була
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Не була
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Оплата
    
    for e in events:
        row_data = [
            e['date'].strftime('%d.%m.%Y %H:%M'),
            e['type'],
            e['details'],
            f"{ e['impact'] if isinstance(e['impact'], str) else ('+' if e['impact'] > 0 else '') + str(e['impact']) }",
            e['amount']
        ]
        ws.append(row_data)
        
        # Отримуємо останній доданий рядок
        last_row = ws.max_row
        
        # Визначаємо колір
        current_fill = None
        if e['type'] == 'Оплата':
            current_fill = fill_blue
        elif e['type'] == 'Заняття':
            if "Присутність" in e['details']:
                current_fill = fill_green
            else:
                current_fill = fill_red
        
        if current_fill:
            for cell in ws[last_row]:
                cell.fill = current_fill
        
        # Додаємо межі для всього рядка
        for cell in ws[last_row]:
            cell.border = thin_border
    
    # Налаштування ширини колонок
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20

    file_path = f"/tmp/report_{student_id}.xlsx"
    wb.save(file_path)
    
    await callback.message.answer_document(
        FSInputFile(file_path, filename=f"Звіт_{student.last_name}_{student.first_name}.xlsx"),
        caption=f"Повний звіт для {student.last_name} {student.first_name}"
    )
    await callback.answer()

# --- БЕКАП ---
@router.message(Command("backup"))
@router.message(F.text == "Бекап бази")
async def cmd_backup(message: Message):
    if str(message.from_user.id) != str(ADMIN_ID):
        await message.answer("У вас немає прав для виконання цієї команди.")
        return

    await message.answer("⏳ Створюю бекап бази даних...")
    
    file_path = "/tmp/backup.sql"
    
    # Отримуємо дані з env
    db_user = os.getenv("DB_USER", os.getenv("POSTGRES_USER", "postgres"))
    db_name = os.getenv("DB_NAME", os.getenv("POSTGRES_DB", "monitoring_db"))
    db_pass = os.getenv("DB_PASS", os.getenv("POSTGRES_PASSWORD", "postgres"))
    db_host = os.getenv("DB_HOST", os.getenv("POSTGRES_HOST", "db"))
    db_type = os.getenv("DB_TYPE", "postgres")
    
    if db_type == "sqlite":
        # Створюємо копію файлу SQLite
        import shutil
        db_file = "monitoring_db.sqlite"
        if os.path.exists(db_file):
            shutil.copy(db_file, file_path)
            # Для SQLite встановлюємо returncode в 0 вручну
            class Process: returncode = 0
            process = Process()
        else:
            await message.answer("❌ Файл бази даних не знайдено.")
            return
    elif db_type == "mysql":
        # Виконуємо mysqldump
        cmd = f"mysqldump -h{db_host} -u{db_user} -p{db_pass} {db_name} > {file_path}"
        process = await asyncio.create_subprocess_shell(cmd)
        await process.wait()
    else:
        # Виконуємо pg_dump
        env = os.environ.copy()
        env["PGPASSWORD"] = db_pass
        process = await asyncio.create_subprocess_exec(
            "pg_dump", "-h", db_host, "-U", db_user, "-f", file_path, db_name,
            env=env
        )
        await process.wait()
    
    if process.returncode == 0:
        await message.answer_document(
            FSInputFile(file_path, filename=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"),
            caption="📦 Повний бекап бази даних успішно створено!"
        )
    else:
        await message.answer("❌ Помилка при створенні бекапу.")
