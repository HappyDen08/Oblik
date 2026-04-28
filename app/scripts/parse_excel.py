import openpyxl
import json
from datetime import datetime
import os

def datetime_serializer(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")

def parse_excel(file_path, output_json):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    data = {
        "transactions": [],
        "attendances": []
    }
    
    # Parse Transactions
    if "транзакції" in wb.sheetnames:
        sheet = wb["транзакції"]
        for row in sheet.iter_rows(values_only=True):
            if not row or row[0] is None: continue
            data["transactions"].append({
                "student_name": row[0],
                "amount": row[1],
                "date": row[2]
            })
            
    # Parse Attendances
    if "відвідування" in wb.sheetnames:
        sheet = wb["відвідування"]
        for row in sheet.iter_rows(values_only=True):
            if not row or row[0] is None: continue
            data["attendances"].append({
                "student_name": row[0],
                "status": row[1],
                "date": row[2]
            })
            
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=datetime_serializer)
    
    print(f"Parsed {len(data['transactions'])} transactions and {len(data['attendances'])} attendances to {output_json}")

if __name__ == "__main__":
    import sys
    input_file = sys.argv[1] if len(sys.argv) > 1 else "Електронна таблиця без назви.xlsx"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "data_dump.json"
    
    if not os.path.exists(input_file):
        # Try finding any xlsx if the name is slightly different
        import glob
        files = glob.glob("*.xlsx")
        if files:
            input_file = files[0]
            
    parse_excel(input_file, output_file)
