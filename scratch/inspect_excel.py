import openpyxl
import json
from datetime import datetime

def datetime_serializer(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} not serializable")

def inspect_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets_info = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
        sheets_info[sheet_name] = rows[:10]  # First 10 rows
        
    print(json.dumps(sheets_info, indent=2, ensure_ascii=False, default=datetime_serializer))

if __name__ == "__main__":
    inspect_excel("./Електронна таблиця без назви.xlsx")
